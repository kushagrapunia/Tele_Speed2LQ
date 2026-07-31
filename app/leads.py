import json
import threading
from datetime import datetime, timezone
from pathlib import Path

from openpyxl import Workbook, load_workbook

from .config import LEADS_EXCEL_PATH

_HEADERS = [
    "Timestamp (UTC)",
    "Chat ID",
    "Telegram Username",
    "Name",
    "Contact",
    "Interest",
    "Preferred Consultation Time",
    "Notes",
]

_leads_path = Path(LEADS_EXCEL_PATH)
_captured_state_path = _leads_path.with_name("captured_leads.json")

# Different chats' leads can now be recorded from separate worker threads at the same time
# (see app/handlers.py) — this serializes access to the shared Excel/JSON files so concurrent
# writes can't silently overwrite each other.
_lock = threading.Lock()


def _ensure_workbook() -> None:
    if not _leads_path.exists():
        _leads_path.parent.mkdir(parents=True, exist_ok=True)
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Leads"
        sheet.append(_HEADERS)
        workbook.save(_leads_path)
        return

    workbook = load_workbook(_leads_path)
    sheet = workbook["Leads"]
    current_header = [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
    if current_header == _HEADERS:
        return

    # File predates a column that's since been added (e.g. "Preferred Consultation Time") —
    # rebuild with it inserted for existing rows rather than silently misaligning or losing data.
    old_rows = list(sheet.iter_rows(min_row=2, values_only=True))
    insert_after = current_header.index("Interest")

    migrated = Workbook()
    migrated_sheet = migrated.active
    migrated_sheet.title = "Leads"
    migrated_sheet.append(_HEADERS)
    for row in old_rows:
        row = list(row)
        row.insert(insert_after + 1, "")
        migrated_sheet.append(row)
    migrated.save(_leads_path)


def _load_captured() -> set[str]:
    if not _captured_state_path.exists():
        return set()
    return set(json.loads(_captured_state_path.read_text(encoding="utf-8")))


def _save_captured(captured: set[str]) -> None:
    _captured_state_path.write_text(json.dumps(sorted(captured)), encoding="utf-8")


def is_lead_captured(chat_id: str) -> bool:
    with _lock:
        return chat_id in _load_captured()


def record_lead(
    chat_id: str,
    telegram_username: str,
    name: str,
    contact: str,
    interest: str,
    preferred_callback_time: str = "",
    notes: str = "",
) -> None:
    with _lock:
        _ensure_workbook()

        workbook = load_workbook(_leads_path)
        sheet = workbook["Leads"]
        sheet.append([
            datetime.now(timezone.utc).isoformat(timespec="seconds"),
            chat_id,
            telegram_username,
            name,
            contact,
            interest,
            preferred_callback_time,
            notes,
        ])
        workbook.save(_leads_path)

        captured = _load_captured()
        captured.add(chat_id)
        _save_captured(captured)
